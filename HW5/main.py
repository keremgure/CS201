#! /usr/bin/python3
import argparse
import sys
import textwrap
import subprocess
import random
import shutil
import time
import os

from openpyxl import Workbook

AVERAGE_N = 10
algTypes = [1,2,3,4]
samples = [1000,10000,100000,500000,"1M","5M","10M"]
canRun = True
def main():
    index = 2
    for algType in algTypes:
        sampleIndex = 0
        for sample in samples:
            global canRun
            canRun = True
            average = 0.0
            ws['A{}'.format(index)] = algType
            ws['B{}'.format(index)] = sample
            for i in range(AVERAGE_N):
                if canRun:
                    duration = run_sample(algType,sample)
                    average = average + duration
                    print('Duration: {}, algType: {}, Curr_Sample: {} ,run#: {}'.format(duration,algType,sample,i))
                else:
                    average = -1
                    print('Not Running older tests reported timeout.')
                    break
            average = average / AVERAGE_N if average != -1 else -1
            print('Average Duration: {} for {} trials for Sample:{} ran by algType: {}'.format(average,AVERAGE_N,sample,algType))
            ws['C{}'.format(index)] = average
            sampleIndex = sampleIndex +1
            index = index + 1
    try:
        shutil.rmtree("./genTests")
    except OSError as e:
        print("Error %s-%s" %(e.filename,e.strerror))
    wb.save(time.strftime("%Y%m%d-%H%M%S")+".xlsx")


def gen_sample(algType,size):
    filename = './genTests/test{}_{}.txt'.format(algType,str(size))
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename,"wb") as curr_sample:
        subprocess.run([testGenFile,str(algType),str(size/2),str(size),str(random.randint(0,size))], stdout=curr_sample)
    curr_sample.close()
    return os.path.relpath(filename)


def run_sample(algType,sample):
    #Format the sample file name to fit in the code.
    if(sample == "1M"):
        sample = 1000000
    elif(sample == "5M"):
        sample = 5000000
    elif(sample == "10M"):
        sample = 10000000
    curr_sample = gen_sample(algType,sample)
    # print('Run Statement: ' + curr_sample)
    try:
        test_case = subprocess.run([runFile,curr_sample],check=False, timeout=timeoutTrigger, stdout=subprocess.PIPE)
        out = (str(test_case.stdout)).split(":")
        end = -5 if sys.platform == 'win32' else -3
        duration = float(format(float(out[2][1:end]),'.8f'))
        return duration
    except (TimeoutError, subprocess.TimeoutExpired) as te:
        # print(te)
        global canRun
        canRun = False
        print('canRun has been set to False due to timeOut')
        return -1



def argsDefiner():
    global parser
    parser = argparse.ArgumentParser(prog='HW5 Test Runner',
      formatter_class=argparse.RawDescriptionHelpFormatter,
      epilog=textwrap.dedent('''\
         additional information:
             Please specify the full path as STRING!
             (use "__")
             for the arguments or put the files
             in the same directory with the script.

         example:
             testRunner.py --use "C:\some_path\\another_path\HW5.exe" --testGen "D:\some_path\\another_path\\testGen.exe -t 60"
         '''))

    parser.add_argument('--use',required=True, help='Specify the program will be used for testing the algorithms.')
    parser.add_argument('--testGen',required=True, help='Specify the program will be used for generating samples.')
    parser.add_argument('-t',type=int, required=False, help='''Specify the time in seconds for a run trial to be terminated after
                                                  the specified time has passed.''')


def argsParser():
    if len(sys.argv) <5:
        parser.print_help(sys.stderr)
        sys.exit(1)

    global runFile
    global testGenFile
    global timeoutTrigger
    args = parser.parse_args()
    timeoutTrigger = 600 if args.t == None else int(args.t)
    runFile = str(args.use)
    testGenFile = str(args.testGen)


if __name__ == "__main__":
    argsDefiner()
    argsParser()
    global wb
    wb = Workbook()
    global ws
    ws = wb.active
    ws['A1'] = "Algorithm Type"
    ws['B1'] = "Sample"
    ws['C1'] = "Duration"
    main()
